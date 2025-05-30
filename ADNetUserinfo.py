import subprocess
import re
import sys
import os
import concurrent.futures
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time

RESUME_FILE = "resume_point.txt"
BATCH_SIZE = 500  # Process users in batches
MAX_WORKERS = 5  # Limit threads
TIMEOUT = 10  # Timeout for subprocess call
RETRY_DELAY = 2  # Delay before retrying failed users


def split_date_time(value):
    if value and value != 'Never':
        try:
            date_part, time_part = value.split(' ', 1)
            return date_part, time_part
        except ValueError:
            return value, ""
    return value, ""


def get_user_info(username):
    try:
        result = subprocess.run(["net", "user", username, "/domain"], capture_output=True, text=True, shell=True, timeout=TIMEOUT)
        output = result.stdout

        if "The command completed successfully" not in output:
            return None

        user_info = {
            "Username": username,
            "Full Name": None,
            "Account Active": None,
            "Password Last Set Date": None,
            "Password Last Set Time": None,
            "Password Expires Date": None,
            "Password Expires Time": None,
            "Password Required": None,
            "Last Logon": None,
            "Logon Script": None,
            "Comment": None
        }

        patterns = {
            "Full Name": r"Full Name\s+(.+)",
            "Account Active": r"Account active\s+(.+)",
            "Password Last Set": r"Password last set\s+(.+)",
            "Password Expires": r"Password expires\s+(.+)",
            "Password Required": r"Password required\s+(.+)",
            "Last Logon": r"Last logon\s+(.+)",
            "Logon Script": r"Logon script\s+(.+)",
            "Comment": r"Comment\s+(.+)"
        }

        for key, pattern in patterns.items():
            match = re.search(pattern, output)
            if match:
                value = match.group(1).strip()
                if key == "Password Last Set":
                    user_info["Password Last Set Date"], user_info["Password Last Set Time"] = split_date_time(value)
                elif key == "Password Expires":
                    user_info["Password Expires Date"], user_info["Password Expires Time"] = split_date_time(value)
                else:
                    user_info[key] = value

        return user_info
    except subprocess.TimeoutExpired:
        with open("error_log.txt", "a") as log_file:
            log_file.write(f"Timeout fetching data for {username}\n")
        return None
    except Exception as e:
        with open("error_log.txt", "a") as log_file:
            log_file.write(f"Error fetching data for {username}: {e}\n")
        return None


def save_to_excel(data, filename):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Full Name", "Account Active", "Password Last Set Date", "Password Last Set Time", "Password Expires Date", "Password Expires Time", "Password Required", "Last Logon", "Logon Script", "Comment"])

    for user_info in data:
        sheet.append([user_info.get("Username"), user_info.get("Full Name"), user_info.get("Account Active"),
                      user_info.get("Password Last Set Date"), user_info.get("Password Last Set Time"),
                      user_info.get("Password Expires Date"), user_info.get("Password Expires Time"),
                      user_info.get("Password Required"), user_info.get("Last Logon"),
                      user_info.get("Logon Script"), user_info.get("Comment")])

    workbook.save(filename)


def process_users(users, output_file):
    start_index = 0

    if os.path.exists(RESUME_FILE):
        with open(RESUME_FILE, "r") as f:
            start_index = int(f.read().strip())

    users = users[start_index:]

    for i in range(0, len(users), BATCH_SIZE):
        batch = users[i:i+BATCH_SIZE]
        all_users_info = []

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            with tqdm(total=len(batch), desc="Processing Batch", unit="user") as pbar:
                futures = {executor.submit(get_user_info, user): user for user in batch}
                for index, future in enumerate(concurrent.futures.as_completed(futures), start=start_index):
                    user_info = future.result()
                    if user_info:
                        all_users_info.append(user_info)
                    pbar.update(1)

                    with open(RESUME_FILE, "w") as f:
                        f.write(str(index + 1))

        if all_users_info:
            save_to_excel(all_users_info, output_file)

        print(f"Batch {i//BATCH_SIZE + 1} processed. Pausing briefly...")
        time.sleep(RETRY_DELAY)

    if os.path.exists(RESUME_FILE):
        os.remove(RESUME_FILE)


def main():
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print("Usage: python script.py users.txt [output_file]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) == 3 else "user_details.xlsx"

    if not output_file.endswith(".xlsx"):
        output_file += ".xlsx"

    try:
        with open(input_file, "r") as file:
            users = [line.strip() for line in file if line.strip()]
    except FileNotFoundError:
        print("Error: Input file not found.")
        sys.exit(1)

    print(f"Total users to process: {len(users)}")
    process_users(users, output_file)
    print(f"User details saved to {output_file}")

if __name__ == "__main__":
    main()
